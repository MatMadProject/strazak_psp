import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

class FirefighterExcelService:
    """
    Serwis do obsługi importu/eksportu strażaków z/do Excel
    """
    
    # Nagłówki zgodne z wymaganiami
    HEADERS = [
        "Id",
        "Imię",
        "Nazwisko",
        "Stopień",
        "Stanowisko",
        "Jednostka Organizacyjna"
    ]
    
    # Mapowanie nagłówków na pola modelu
    COLUMN_MAPPING = {
        "Imię": "imie",
        "Nazwisko": "nazwisko",
        "Stopień": "stopien",
        "Stanowisko": "stanowisko",
        "Jednostka Organizacyjna": "jednostka"
    }
    
    def validate_file(self, file_path: Path) -> tuple[bool, str]:
        """
        Walidacja pliku Excel
        Returns: (is_valid, error_message)
        """
        try:
            if not file_path.exists():
                return False, "Plik nie istnieje"
            
            if file_path.suffix.lower() not in ['.xlsx']:
                return False, "Nieprawidłowe rozszerzenie pliku (wymagany .xlsx)"
            
            # Sprawdź czy plik można otworzyć
            try:
                df = pd.read_excel(file_path)
                if df.empty:
                    return False, "Plik Excel jest pusty"
                
                # Sprawdź nagłówki
                missing_headers = []
                for header in self.HEADERS[1:]:  # Pomijamy "Id" bo jest opcjonalne
                    if header not in df.columns:
                        missing_headers.append(header)
                
                if missing_headers:
                    return False, f"Brakujące kolumny: {', '.join(missing_headers)}"
                
            except Exception as e:
                return False, f"Nie można odczytać pliku Excel: {str(e)}"
            
            return True, ""
        except Exception as e:
            return False, f"Błąd walidacji: {str(e)}"
    
    def process_excel_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """
        Przetwarza plik Excel i zwraca listę słowników ze strażakami
        
        Returns: Lista słowników z danymi strażaków
        """
        import traceback
        
        try:
            print(f"[FIREFIGHTER EXCEL] Rozpoczynam przetwarzanie: {file_path}")
            
            # Walidacja
            is_valid, error = self.validate_file(file_path)
            if not is_valid:
                print(f"[ERROR] [FIREFIGHTER EXCEL] Walidacja nieudana: {error}")
                raise ValueError(error)
            
            # Odczyt pliku
            print(f"[FIREFIGHTER EXCEL] Odczyt pliku...")
            df = pd.read_excel(file_path)
            print(f"[FIREFIGHTER EXCEL] Odczytano {len(df)} wierszy")
            
            # Przetwarzanie rekordów
            firefighters = []
            skipped_rows = 0
            
            for idx, row in df.iterrows():
                try:
                    firefighter = self._parse_row_to_firefighter(row)
                    if firefighter:
                        firefighters.append(firefighter)
                    else:
                        skipped_rows += 1
                except Exception as e:
                    print(f"[WARN]] [FIREFIGHTER EXCEL] Błąd w wierszu {idx + 2}: {e}")
                    skipped_rows += 1
                    continue
            
            print(f"[INFO] [FIREFIGHTER EXCEL] Przetworzono {len(firefighters)} strażaków")
            if skipped_rows > 0:
                print(f"[WARN] [FIREFIGHTER EXCEL] Pominięto {skipped_rows} wierszy")
            
            return firefighters
            
        except Exception as e:
            print(f"[ERROR] [FIREFIGHTER EXCEL] BŁĄD: {str(e)}")
            traceback.print_exc()
            raise Exception(f"Błąd przetwarzania pliku: {str(e)}")
    
    def _parse_row_to_firefighter(self, row: pd.Series) -> Dict[str, Any]:
        """
        Parsuje wiersz pandas Series do słownika dla modelu Firefighter
        """
        try:
            # Pobierz imię i nazwisko
            imie = str(row.get('Imię', '')).strip()
            nazwisko = str(row.get('Nazwisko', '')).strip()
            
            # Walidacja wymaganych pól
            if not imie or not nazwisko:
                print(f"[WARN] Pominięto wiersz: brak imienia lub nazwiska")
                return None
            
            # Połącz nazwisko i imię zgodnie z modelem
            nazwisko_imie = f"{nazwisko.upper()} {imie}"
            
            stopien = str(row.get('Stopień', '')).strip()
            stanowisko = str(row.get('Stanowisko', '')).strip()
            jednostka = str(row.get('Jednostka Organizacyjna', '')).strip()
            
            # Walidacja wymaganych pól
            if not stopien or not stanowisko or not jednostka:
                print(f"[WARN] Pominięto wiersz {nazwisko_imie}: brakuje wymaganych danych")
                return None
            
            firefighter = {
                "nazwisko_imie": nazwisko_imie,
                "stopien": stopien,
                "stanowisko": stanowisko,
                "jednostka": jednostka
            }
            
            return firefighter
            
        except Exception as e:
            print(f"[ERROR] Błąd parsowania wiersza: {e}")
            return None
    
    def create_template_file(self) -> BytesIO:
        """
        Tworzy pusty szablon pliku Excel z przykładowymi danymi
        
        Returns: BytesIO z plikiem Excel
        """
        # Utwórz workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Strażacy"
        
        # Style nagłówków
        #header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Arial", bold=True, size=10)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Dodaj nagłówki
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            #cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Przykładowe dane (3 przykładowe wiersze)
        example_data = [
            [1, "Jan", "KOWALSKI", "Kapitan", "Dowódca zmiany", "KP PSP Kraków"],
            [2, "Anna", "NOWAK", "Aspirant", "Ratownik", "KP PSP Kraków"],
            [3, "Piotr", "WIŚNIEWSKI", "Starszy strażak", "Kierowca", "KM PSP Zakopane"]
        ]
        
        for row_idx, data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name="Arial", size=10)
        
        # Dodaj instrukcje
        instruction_row = 2
        ws.cell(row=instruction_row, column=8, value="INSTRUKCJA:")
        ws.cell(row=instruction_row, column=8).font = Font(bold=True, size=11)
        
        instructions = [
            "1. Usuń przykładowe wiersze (2-4)",
            "2. Wypełnij dane strażaków w kolejnych wierszach",
            "3. Kolumna 'Id' jest opcjonalna - możesz ją zostawić pustą",
            "4. Wszystkie pozostałe kolumny są wymagane",
            "5. Zapisz plik i zaimportuj w aplikacji"
        ]
        
        for i, instruction in enumerate(instructions, start=1):
            ws.cell(row=instruction_row + i, column=8, value=instruction)
        
        # Dostosuj szerokość kolumn
        column_widths = [8, 15, 20, 20, 25, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # Zapisz do BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def get_file_summary(self, file_path: Path) -> Dict[str, Any]:
        """Zwraca podsumowanie pliku Excel"""
        try:
            df = pd.read_excel(file_path)
            return {
                "rows_count": len(df),
                "columns_count": len(df.columns),
                "columns": df.columns.tolist(),
                "preview": df.head(5).to_dict('records')
            }
        except Exception as e:
            return {"error": str(e)}
        
    def export_to_excel(self, firefighters: List[Dict[str, Any]]) -> BytesIO:
        """
        Eksportuje listę strażaków do pliku Excel
        
        Args:
            firefighters: Lista słowników z danymi strażaków
        
        Returns: BytesIO z plikiem Excel
        """
        # Utwórz workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Strażacy"
        
        # Style nagłówków
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Dodaj nagłówki
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Dodaj dane strażaków
        for row_idx, firefighter in enumerate(firefighters, start=2):
            # Rozdziel nazwisko_imie na nazwisko i imię
            nazwisko_imie = firefighter.get('nazwisko_imie', '')
            parts = nazwisko_imie.split(' ', 1)
            nazwisko = parts[0] if len(parts) > 0 else ''
            imie = parts[1] if len(parts) > 1 else ''
            
            # Wypełnij wiersz
            row_data = [
                firefighter.get('id', row_idx - 1),
                imie,
                nazwisko,
                firefighter.get('stopien', ''),
                firefighter.get('stanowisko', ''),
                firefighter.get('jednostka', '')
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name="Arial", size=10)
        
        # Dostosuj szerokość kolumn
        column_widths = [8, 15, 20, 20, 25, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # Dodaj informacje o eksporcie
        info_row = len(firefighters) + 3
        ws.cell(row=info_row, column=1, value=f"Wyeksportowano: {len(firefighters)} strażaków")
        ws.cell(row=info_row, column=1).font = Font(italic=True, size=9, color="666666")
        
        from datetime import datetime
        ws.cell(row=info_row + 1, column=1, value=f"Data eksportu: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=info_row + 1, column=1).font = Font(italic=True, size=9, color="666666")

        # ws.cell(row=info_row + 2, column=1, value=f"Wyeksportowany przez: {self.user_service.get_current_user().get('name')} {self.user_service.get_current_user().get('last_name')}")
        # ws.cell(row=info_row + 2, column=1).font = Font(italic=True, size=9, color="666666")
        
        ws.cell(row=info_row + 2, column=1, value=f"app.straznica.com.pl 2026 wszelkie prawa zastrzeżone")
        ws.cell(row=info_row + 2, column=1).font = Font(italic=True, size=9, color="666666")
        # Zapisz do BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def export_to_csv(self, firefighters: List[Dict[str, Any]]) -> BytesIO:
        """
        Eksportuje listę strażaków do pliku CSV
        
        Args:
            firefighters: Lista słowników z danymi strażaków
        
        Returns: BytesIO z plikiem CSV
        """
        # Przygotuj dane do DataFrame
        data = []
        for firefighter in firefighters:
            # Rozdziel nazwisko_imie na nazwisko i imię
            nazwisko_imie = firefighter.get('nazwisko_imie', '')
            parts = nazwisko_imie.split(' ', 1)
            nazwisko = parts[0] if len(parts) > 0 else ''
            imie = parts[1] if len(parts) > 1 else ''
            
            data.append({
                'Id': firefighter.get('id', ''),
                'Imię': imie,
                'Nazwisko': nazwisko,
                'Stopień': firefighter.get('stopien', ''),
                'Stanowisko': firefighter.get('stanowisko', ''),
                'Jednostka Organizacyjna': firefighter.get('jednostka', '')
            })
        
        # Utwórz DataFrame
        df = pd.DataFrame(data)
        
        # Zapisz do BytesIO
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')  # utf-8-sig dla polskich znaków w Excel
        output.seek(0)
        
        return output    