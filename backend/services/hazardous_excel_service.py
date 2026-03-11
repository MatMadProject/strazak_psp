"""
backend/services/hazardous_excel_service.py

Eksport rekordów Dodatku Szkodliwego do Excel i CSV.
Wzorowany na HazardousDegreesExcelService — identyczny styl i struktura.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from typing import List, Dict, Any
from datetime import datetime
import csv


class HazardousExcelService:
    """Serwis do eksportu rekordów Dodatku Szkodliwego do Excel/CSV"""

    HEADERS = [
        "Jednostka",
        "Nazwisko i imię",
        "Funkcja",
        "Nr meldunku",
        "Czas od",
        "Czas do",
        "Czas udziału",
        "P",
        "MZ",
        "AF",
        "Dodatek szkodliwy",
        "Stopień szkodl. (plik)",
        "Przypisany stopień",
    ]

    COL_WIDTHS = [16, 28, 18, 16, 20, 20, 12, 5, 5, 5, 18, 22, 40]

    def _record_to_row(self, record: Dict[str, Any]) -> list:
        """Mapuje rekord do wiersza Excel/CSV"""
        degree = record.get("hazardous_degree")
        if degree:
            degree_label = f"{degree.get('stopien','')}.{degree.get('punkt','')} — {degree.get('opis','')}"
        else:
            degree_label = ""

        return [
            record.get("jednostka")            or "",
            record.get("nazwisko_imie")         or "",
            record.get("funkcja")               or "",
            record.get("nr_meldunku")           or "",
            record.get("czas_od")               or "",
            record.get("czas_do")               or "",
            record.get("czas_udzialu")          or "",
            record.get("p")                     or "",
            record.get("mz")                    or "",
            record.get("af")                    or "",
            record.get("dodatek_szkodliwy")     or "",
            record.get("stopien_szkodliwosci")  or "",
            degree_label,
        ]

    # ── Eksport XLSX ─────────────────────────────────────────────────────────

    def export_to_excel(self, records: List[Dict[str, Any]]) -> BytesIO:
        """Eksportuje listę rekordów do pliku Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dodatek Szkodliwy"

        # Style nagłówka — identyczne jak HazardousDegreesExcelService
        header_fill      = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        header_font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border      = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        # Nagłówki
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell            = ws.cell(row=1, column=col_idx, value=header)
            cell.fill       = header_fill
            cell.font       = header_font
            cell.alignment  = header_alignment
            cell.border     = thin_border

        ws.row_dimensions[1].height = 35

        # Dane
        for row_idx, record in enumerate(records, start=2):
            row_data = self._record_to_row(record)
            for col_idx, value in enumerate(row_data, start=1):
                cell           = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font      = Font(name="Arial", size=10)
                cell.border    = thin_border

        # Szerokości kolumn
        for col_idx, width in enumerate(self.COL_WIDTHS, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

        # Stopka — identyczna jak HazardousDegreesExcelService
        info_row = len(records) + 3
        ws.cell(row=info_row,     column=1, value=f"Wyeksportowano: {len(records)} rekordów").font = Font(italic=True, size=9, color="666666")
        ws.cell(row=info_row + 1, column=1, value=f"Data eksportu: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(italic=True, size=9, color="666666")
        ws.cell(row=info_row + 2, column=1, value="MatMad Software 2026 wszelkie prawa zastrzeżone").font = Font(italic=True, size=9, color="666666")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ── Eksport CSV ──────────────────────────────────────────────────────────

    def export_to_csv(self, records: List[Dict[str, Any]]) -> BytesIO:
        """Eksportuje listę rekordów do pliku CSV"""
        from io import StringIO
        output = StringIO()
        writer = csv.writer(output, delimiter=";")

        writer.writerow(self.HEADERS)
        for record in records:
            writer.writerow(self._record_to_row(record))

        result = BytesIO(output.getvalue().encode("utf-8-sig"))  # utf-8-sig dla polskich znaków w Excel
        result.seek(0)
        return result