import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

class DeparturesExcelService:
    """
    Serwis do obsługi eksportu wyjazdów do Excel/CSV
    """
    
    # Nagłówki zgodne z wymaganiami
    HEADERS = [
        "Id",
        "Imię",
        "Nazwisko",
        "Funkcja",
        "Nr meldunku",
        "Data zdarzenia",
        "Rodzaj zdarzenia",
        "Zaliczono do 0,5%"
    ]
    
    def _get_event_type(self, record: Dict[str, Any]) -> str:
        """
        Określa rodzaj zdarzenia na podstawie kolumn P, MZ, AF
        """
        if str(record.get('p', '')).strip() == '1':
            return 'P'
        elif str(record.get('mz', '')).strip() == '1':
            return 'MZ'
        elif str(record.get('af', '')).strip() == '1':
            return 'AF'
        return ''
    
    def _parse_nazwisko_imie(self, nazwisko_imie: str) -> tuple:
        """
        Rozdziela 'KOWALSKI Jan' na nazwisko i imię
        """
        parts = nazwisko_imie.strip().split(' ', 1)
        nazwisko = parts[0] if len(parts) > 0 else ''
        imie = parts[1] if len(parts) > 1 else ''
        return nazwisko, imie
    
    def export_to_excel(self, records: List[Dict[str, Any]]) -> BytesIO:
        """
        Eksportuje listę wyjazdów do pliku Excel
        
        Args:
            records: Lista słowników z danymi wyjazdów
        
        Returns: BytesIO z plikiem Excel
        """
        # Utwórz workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Wyjazdy"
        
        # Style nagłówków
        header_font = Font(name="Arial", bold=True, size=10)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Dodaj nagłówki
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Dodaj dane wyjazdów
        for row_idx, record in enumerate(records, start=2):
            # Rozdziel nazwisko_imie
            nazwisko, imie = self._parse_nazwisko_imie(record.get('nazwisko_imie', ''))
            
            # Określ rodzaj zdarzenia
            rodzaj_zdarzenia = self._get_event_type(record)
            
            # Konwertuj zaliczono_do_emerytury
            zaliczono = record.get('zaliczono_do_emerytury', '')
            zaliczono_text = 'Tak' if str(zaliczono) == '1' else 'Nie' if str(zaliczono) == '0' else ''
            
            # Wypełnij wiersz
            row_data = [
                record.get('id', row_idx - 1),
                imie,
                nazwisko,
                record.get('funkcja', ''),
                record.get('nr_meldunku', ''),
                record.get('czas_rozp_zdarzenia', ''),
                rodzaj_zdarzenia,
                zaliczono_text
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(name="Arial", size=10)
        
        # Dostosuj szerokość kolumn
        column_widths = [8, 15, 20, 20, 20, 20, 15, 15]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # Dodaj informacje o eksporcie
        info_row = len(records) + 3
        ws.cell(row=info_row, column=1, value=f"Wyeksportowano: {len(records)} wyjazdów")
        ws.cell(row=info_row, column=1).font = Font(italic=True, size=9, color="666666")
        
        from datetime import datetime
        ws.cell(row=info_row + 1, column=1, value=f"Data eksportu: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=info_row + 1, column=1).font = Font(italic=True, size=9, color="666666")
        
        ws.cell(row=info_row + 2, column=1, value=f"app.straznica.com.pl 2026 wszelkie prawa zastrzeżone")
        ws.cell(row=info_row + 2, column=1).font = Font(italic=True, size=9, color="666666")
        
        # Zapisz do BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def export_to_csv(self, records: List[Dict[str, Any]]) -> BytesIO:
        """
        Eksportuje listę wyjazdów do pliku CSV
        
        Args:
            records: Lista słowników z danymi wyjazdów
        
        Returns: BytesIO z plikiem CSV
        """
        # Przygotuj dane do DataFrame
        data = []
        for record in records:
            # Rozdziel nazwisko_imie
            nazwisko, imie = self._parse_nazwisko_imie(record.get('nazwisko_imie', ''))
            
            # Określ rodzaj zdarzenia
            rodzaj_zdarzenia = self._get_event_type(record)
            
            # Konwertuj zaliczono_do_emerytury
            zaliczono = record.get('zaliczono_do_emerytury', '')
            zaliczono_text = 'Tak' if str(zaliczono) == '1' else 'Nie' if str(zaliczono) == '0' else ''
            
            data.append({
                'Id': record.get('id', ''),
                'Imię': imie,
                'Nazwisko': nazwisko,
                'Funkcja': record.get('funkcja', ''),
                'Nr meldunku': record.get('nr_meldunku', ''),
                'Data zdarzenia': record.get('czas_rozp_zdarzenia', ''),
                'Rodzaj zdarzenia': rodzaj_zdarzenia,
                'Zaliczono do 0,5%': zaliczono_text
            })
        
        # Utwórz DataFrame
        df = pd.DataFrame(data)
        
        # Zapisz do BytesIO
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')  # utf-8-sig dla polskich znaków w Excel
        output.seek(0)
        
        return output