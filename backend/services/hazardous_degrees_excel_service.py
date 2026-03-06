"""
backend/services/hazardous_degrees_excel_service.py

Import/eksport stopni szkodliwości z/do Excel i CSV.
Wzorowany na FirefighterExcelService - zachowuje identyczny styl.
"""
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime


class HazardousDegreesExcelService:
    """Serwis do obsługi importu/eksportu stopni szkodliwości z/do Excel"""

    # Nagłówki pliku Excel (import i eksport)
    HEADERS = ["Id", "Stopień", "Punkt", "Opis", "Uwagi"]

    # Mapowanie nagłówków → pola modelu
    COLUMN_MAPPING = {
        "Stopień": "stopien",
        "Punkt":   "punkt",
        "Opis":    "opis",
        "Uwagi":   "uwagi",
    }

    # ── Walidacja ───────────────────────────────────────────────────────────

    def validate_file(self, file_path: Path) -> tuple[bool, str]:
        """Walidacja pliku Excel. Returns: (is_valid, error_message)"""
        try:
            if not file_path.exists():
                return False, "Plik nie istnieje"

            if file_path.suffix.lower() not in [".xlsx"]:
                return False, "Nieprawidłowe rozszerzenie pliku (wymagany .xlsx)"

            df = pd.read_excel(file_path)
            if df.empty:
                return False, "Plik Excel jest pusty"

            # Wymagane kolumny (Id jest opcjonalne)
            required = ["Stopień", "Punkt", "Opis"]
            missing = [h for h in required if h not in df.columns]
            if missing:
                return False, f"Brakujące kolumny: {', '.join(missing)}"

            return True, ""
        except Exception as e:
            return False, f"Błąd walidacji: {str(e)}"

    # ── Import ──────────────────────────────────────────────────────────────

    def process_excel_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """
        Przetwarza plik Excel i zwraca listę słowników gotowych do zapisu w bazie.
        Wzorowane na FirefighterExcelService.process_excel_file.
        """
        import traceback

        try:
            print(f"[HAZARDOUS EXCEL] Rozpoczynam przetwarzanie: {file_path}")

            is_valid, error = self.validate_file(file_path)
            if not is_valid:
                print(f"[ERROR] [HAZARDOUS EXCEL] Walidacja nieudana: {error}")
                raise ValueError(error)

            df = pd.read_excel(file_path)
            print(f"[HAZARDOUS EXCEL] Odczytano {len(df)} wierszy")

            records = []
            skipped = 0

            for idx, row in df.iterrows():
                try:
                    record = self._parse_row(row)
                    if record:
                        records.append(record)
                    else:
                        skipped += 1
                except Exception as e:
                    print(f"[WARN] [HAZARDOUS EXCEL] Błąd w wierszu {idx + 2}: {e}")
                    skipped += 1
                    continue

            print(f"[INFO] [HAZARDOUS EXCEL] Przetworzono {len(records)} rekordów")
            if skipped:
                print(f"[WARN] [HAZARDOUS EXCEL] Pominięto {skipped} wierszy")

            return records

        except Exception as e:
            print(f"[ERROR] [HAZARDOUS EXCEL] BŁĄD: {str(e)}")
            traceback.print_exc()
            raise Exception(f"Błąd przetwarzania pliku: {str(e)}")

    def _parse_row(self, row: pd.Series) -> Dict[str, Any]:
        """Parsuje jeden wiersz DataFrame do słownika dla modelu HazardousDegree"""
        try:
            # Stopień i punkt - konwertuj do int
            stopien_raw = row.get("Stopień", "")
            punkt_raw   = row.get("Punkt", "")

            if pd.isna(stopien_raw) or pd.isna(punkt_raw):
                print(f"[WARN] Pominięto wiersz: brak stopnia lub punktu")
                return None

            try:
                stopien = int(float(str(stopien_raw).strip()))
                punkt   = int(float(str(punkt_raw).strip()))
            except ValueError:
                print(f"[WARN] Pominięto wiersz: stopień/punkt nie są liczbą: {stopien_raw}, {punkt_raw}")
                return None

            opis = str(row.get("Opis", "")).strip()
            if not opis or opis == "nan":
                print(f"[WARN] Pominięto wiersz {stopien}.{punkt}: brak opisu")
                return None

            uwagi_raw = row.get("Uwagi", "")
            uwagi = str(uwagi_raw).strip() if not pd.isna(uwagi_raw) else None
            if uwagi == "nan":
                uwagi = None

            return {
                "stopien": stopien,
                "punkt":   punkt,
                "opis":    opis,
                "uwagi":   uwagi,
            }

        except Exception as e:
            print(f"[ERROR] Błąd parsowania wiersza: {e}")
            return None

    # ── Szablon ─────────────────────────────────────────────────────────────

    def create_template_file(self) -> BytesIO:
        """Tworzy pusty szablon Excel z przykładowymi danymi"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stopnie Szkodliwości"

        header_font      = Font(name="Arial", bold=True, size=10)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.alignment = header_alignment

        example_data = [
            [1, 1, 1, "Praca w hałasie o poziomie 80-85 dB", "Dotyczy stanowisk przy maszynach"],
            [2, 1, 2, "Praca z substancjami drażniącymi", "Wymagane środki ochrony indywidualnej"],
            [3, 2, 1, "Praca w zapyleniu powyżej NDS", "Pomiar co 6 miesięcy"],
        ]

        for row_idx, data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font      = Font(name="Arial", size=10)

        # Instrukcje w osobnej kolumnie
        ws.cell(row=2, column=7, value="INSTRUKCJA:").font = Font(bold=True, size=11)
        instructions = [
            "1. Usuń przykładowe wiersze (2-4)",
            "2. Kolumna 'Id' jest opcjonalna",
            "3. Stopień i Punkt muszą być liczbami całkowitymi",
            "4. Kolumna 'Opis' jest wymagana",
            "5. Zapisz plik i zaimportuj w aplikacji",
        ]
        for i, instr in enumerate(instructions, start=1):
            ws.cell(row=2 + i, column=7, value=instr)

        # Szerokości kolumn
        for col_idx, width in enumerate([8, 10, 10, 60, 40], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ── Eksport XLSX ────────────────────────────────────────────────────────

    def export_to_excel(self, records: List[Dict[str, Any]]) -> BytesIO:
        """Eksportuje listę rekordów do pliku Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stopnie Szkodliwości"

        header_fill      = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_alignment

        for row_idx, record in enumerate(records, start=2):
            row_data = [
                record.get("id", row_idx - 1),
                record.get("stopien", ""),
                record.get("punkt", ""),
                record.get("opis", ""),
                record.get("uwagi", ""),
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font      = Font(name="Arial", size=10)

        # Szerokości kolumn
        for col_idx, width in enumerate([8, 10, 10, 60, 40], start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Stopka
        info_row = len(records) + 3
        ws.cell(row=info_row,     column=1, value=f"Wyeksportowano: {len(records)} rekordów").font = Font(italic=True, size=9, color="666666")
        ws.cell(row=info_row + 1, column=1, value=f"Data eksportu: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(italic=True, size=9, color="666666")
        ws.cell(row=info_row + 2, column=1, value="MatMad Software 2026 wszelkie prawa zastrzeżone").font = Font(italic=True, size=9, color="666666")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ── Eksport CSV ─────────────────────────────────────────────────────────

    def export_to_csv(self, records: List[Dict[str, Any]]) -> BytesIO:
        """Eksportuje listę rekordów do pliku CSV"""
        data = [
            {
                "Id":      r.get("id", ""),
                "Stopień": r.get("stopien", ""),
                "Punkt":   r.get("punkt", ""),
                "Opis":    r.get("opis", ""),
                "Uwagi":   r.get("uwagi", ""),
            }
            for r in records
        ]

        df = pd.DataFrame(data)
        output = BytesIO()
        df.to_csv(output, index=False, encoding="utf-8-sig")  # utf-8-sig dla polskich znaków w Excel
        output.seek(0)
        return output